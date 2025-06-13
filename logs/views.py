from django.shortcuts import render, redirect, get_object_or_404
from .models import Student, LateLog
import matplotlib.pyplot as plt
import plotly.graph_objects as go
import plotly.express as px
import pandas as pd
import io
from .forms import StudentForm
from django.http import HttpResponse, JsonResponse
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from django.utils import timezone
from django.db.models import Q
from datetime import datetime
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.utils.timezone import localtime
import base64
import pytz
import random
from django.shortcuts import render
from django.db.models import Q
from .models import Student
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from .forms import LoginForm
from django.utils.timezone import localtime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Q
from django.shortcuts import render, redirect
from .models import Student
from django.db.models import Q
import pandas as pd
from django.http import HttpResponse
from .forms import UserRegistrationForm
from django.contrib.auth.models import User, Group
from .models import Profile
from .forms import UserProfileEditForm
from django.contrib.auth.models import Group
from django.http import HttpResponseForbidden
from django.http import HttpResponseForbidden
from functools import wraps

# Custom decorator to restrict access based on user group
def group_required(group_name):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if request.user.groups.filter(name=group_name).exists():
                return view_func(request, *args, **kwargs)
            else:
                return HttpResponseForbidden("You do not have permission to access this page.")
        return _wrapped_view
    return decorator

def home(request):
    # Retrieve late logs and prepare data for the graph
    late_logs = LateLog.objects.select_related('student')
    data = pd.DataFrame(list(late_logs.values(
        'student__year_level',
        'student__first_name',
        'student__last_name',
        'student__learner_reference_number',
        'late_minutes',
        'scan_time',
        'student__address',
        'student__sex'
    )))

    # Group by year level
    year_level_data = data.groupby('student__year_level')['student__first_name'].count().reset_index()
    year_level_data['student__year_level'] = year_level_data['student__year_level'].astype(str)

    # Create the bar chart for late entries by year level
    year_level_chart = go.Figure(
        data=[go.Bar(
            x=year_level_data['student__year_level'],
            y=year_level_data['student__first_name'],
            marker=dict(color='blue')
        )]
    )
    year_level_chart.update_layout(
        title="Late Entries by Year Level",
        xaxis_title="Year Level",
        yaxis_title="Number of Late Entries"
    )

    # Pass the chart as an HTML string to the template
    year_level_chart_html = year_level_chart.to_html(full_html=False)

    # Render the home page with the chart
    return render(request, 'home.html', {'year_level_chart': year_level_chart_html})


@login_required
def index(request):
    # Check if the user is a student or a teacher
    is_student = request.user.groups.filter(name='student').exists()
    is_teacher = request.user.groups.filter(name='teacher').exists()

    # Pass this information to the template
    return render(request, 'logs/index.html', {
        'is_student': is_student,
        'is_teacher': is_teacher,
    })

@login_required
@group_required('teacher')
def register_user(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            # Get cleaned data from the form
            username = form.cleaned_data['username']
            email = form.cleaned_data['email']
            password = form.cleaned_data['password1']
            user_type = form.cleaned_data['user_type']

            # Create the user
            user = User.objects.create_user(username=username, email=email, password=password)

            # Assign group based on user_type (either student or teacher)
            if user_type == 'student':
                group = Group.objects.get(name='student')  # Ensure 'student' group exists
            else:
                group = Group.objects.get(name='teacher')  # Ensure 'teacher' group exists

            user.groups.add(group)  # Add the user to the respective group
            user.save()

            # Log the user in after registration
            login(request, user)
            messages.success(request, f'User {username} has been successfully registered and logged in.')

            # Redirect to the user profile list
            return redirect('view_user_profile')  # Redirect to the user profile page after successful registration
        else:
            messages.error(request, 'There was an error with your form. Please check your input.')
    else:
        form = UserRegistrationForm()

    return render(request, 'logs/register_user.html', {'form': form})


def home(request):
    return redirect('login')  # Redirect root URL to the login page

@login_required
@group_required('teacher')
def view_user_profile(request):
    # Get the search query from the request
    search_query = request.GET.get('search', '')

    # Filter users based on the search query for multiple fields (username, first_name, last_name, email)
    if search_query:
        users = User.objects.filter(
            Q(username__icontains=search_query) |  # Search by username
            Q(first_name__icontains=search_query) |  # Search by first name
            Q(last_name__icontains=search_query) |  # Search by last name
            Q(email__icontains=search_query)  # Search by email
        )
    else:
        users = User.objects.all()  # If no search query, show all users

    # Combine first_name and last_name for each user as full name
    for user in users:
        user.full_name = f"{user.first_name} {user.last_name}"

    return render(request, 'logs/view_user_profile.html', {'users': users, 'search_query': search_query})


@login_required
@group_required('teacher')
def edit_user_profile(request, user_id):
    # Get the user object or return 404 if not found
    user = get_object_or_404(User, pk=user_id)

    # Ensure the user has a profile; if not, create one
    if not hasattr(user, 'profile'):
        Profile.objects.create(user=user)

    if request.method == 'POST':
        form = UserProfileEditForm(request.POST, instance=user)
        if form.is_valid():
            # Save the user fields (username, email)
            user = form.save()

            # Manually update the user_type in the Profile model
            profile = user.profile  # Access the Profile related to the user
            profile.user_type = form.cleaned_data['user_type']  # Update user_type from the form
            profile.save()  # Save the Profile model

            messages.success(request, f'User {user.username} has been successfully updated.')
            return redirect('view_user_profile')  # Redirect after successful update
    else:
        form = UserProfileEditForm(instance=user)

    return render(request, 'logs/edit_user_profile.html', {'form': form, 'user': user})


def chart_reports(request):
    # Get filters from request
    selected_year_level = request.GET.get('year_level', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    address_filter = request.GET.get('address', '')
    sex_filter = request.GET.get('sex', '')

    # Initialize filters for the title
    filters = []

    # Append year level filter to the title suffix
    if selected_year_level:
        filters.append(f"Grade {selected_year_level}")

    # Append date filter to the title suffix if start and end dates are provided
    if start_date and end_date:
        filters.append(f"From {datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y')} To {datetime.strptime(end_date, '%Y-%m-%d').strftime('%m/%d/%Y')}")

    # Append address filter to the title suffix if provided
    if address_filter:
        filters.append(address_filter)

    # Append sex filter to the title suffix if provided
    if sex_filter == 'M':
        filters.append("Male")
    elif sex_filter == 'F':
        filters.append("Female")

    title_suffix = ', '.join(filters)

    # Retrieve late logs and prepare data
    late_logs = LateLog.objects.select_related('student')
    data = pd.DataFrame(list(late_logs.values(
        'student__year_level',
        'student__section',
        'student__first_name',
        'student__last_name',
        'student__learner_reference_number',
        'late_minutes',
        'scan_time',
        'student__address',
        'student__sex'
    )))

    # Convert scan_time to local time
    data['scan_time'] = pd.to_datetime(data['scan_time'], errors='coerce')
    data['scan_time'] = data['scan_time'].apply(localtime)

    # Apply date filters if available
    if start_date and end_date:
        start_date = pd.to_datetime(start_date)
        end_date = pd.to_datetime(end_date)
        utc = pytz.UTC
        start_date = utc.localize(start_date)
        end_date = utc.localize(end_date)

        if start_date.date() == end_date.date():
            data = data[data['scan_time'].dt.date == start_date.date()]
        else:
            data = data[(data['scan_time'] >= start_date) & (data['scan_time'] <= end_date)]

    # Filter data by Year Level if selected
    if selected_year_level:
        data = data[data['student__year_level'] == int(selected_year_level)]

    # Filter by address (barangay) if selected
    if address_filter:
        data = data[data['student__address'].str.contains(address_filter, na=False, case=False)]

    # Filter by sex if selected
    if sex_filter:
        data = data[data['student__sex'] == sex_filter]

    # Prepare data for bar charts
    fixed_colors = {
        7: 'green',
        8: 'yellow',
        9: 'red',
        10: 'blue',
        11: 'pink',
        12: 'violet'
    }

    # === Year Level Chart ===
    year_level_data = data.groupby('student__year_level')['student__first_name'].count().reset_index()
    year_level_data['student__year_level'] = year_level_data['student__year_level'].astype(str)

    year_level_chart = go.Figure(
        data=[go.Bar(
            x=year_level_data['student__year_level'],
            y=year_level_data['student__first_name'],
            marker=dict(color=[fixed_colors.get(int(x), 'gray') for x in year_level_data['student__year_level']])
        )]
    )
    year_level_chart.update_layout(
        title="Late Entries by Year Level",
        xaxis_title="Year Level",
        yaxis_title="Number of Late Entries"
    )
    year_level_chart_html = year_level_chart.to_html(full_html=False)

    # === Section Chart ===
    section_group = data.groupby('student__section')['student__first_name'].count().sort_values(ascending=False)
    section_data_all = section_group.reset_index()
    section_data_all.columns = ['section', 'late_count']

    section_colors = {}
    for section in section_data_all['section']:
        section_years = data[data['student__section'] == section]['student__year_level']
        dominant_year = section_years.mode()[0] if not section_years.empty else None
        section_colors[section] = fixed_colors.get(dominant_year, 'gray')

    section_bars = [
        go.Bar(
            x=[row['section']],
            y=[row['late_count']],
            marker_color=section_colors.get(row['section'], 'gray'),
            name=row['section']
        )
        for _, row in section_data_all.iterrows()
    ]

    section_chart = go.Figure(data=section_bars)
    section_chart.update_layout(
        title="Late Entries by Section",
        xaxis_title="Section",
        yaxis_title="Number of Late Entries",
        xaxis=dict(range=[-0.5, 9.5] if len(section_data_all) > 10 else None),
        showlegend=False
    )
    section_chart_html = section_chart.to_html(full_html=False)

    # === Student Chart ===
    data['student_name'] = data['student__first_name'] + ' ' + data['student__last_name'] + ' (' + data['student__learner_reference_number'].astype(str) + ')'
    student_data_all = data.groupby('student_name').agg({
        'late_minutes': 'count',
        'student__year_level': 'first'
    }).reset_index().sort_values(by='late_minutes', ascending=False)

    student_colors = {
        row['student_name']: fixed_colors.get(row['student__year_level'], 'gray')
        for _, row in student_data_all.iterrows()
    }

    student_bars = [
        go.Bar(
            x=[row['student_name']],
            y=[row['late_minutes']],
            marker_color=student_colors.get(row['student_name'], 'gray'),
            name=row['student_name']
        )
        for _, row in student_data_all.iterrows()
    ]

    student_chart = go.Figure(data=student_bars)
    student_chart.update_layout(
        title="Late Entries by Student",
        xaxis_title="Student",
        yaxis_title="Number of Late Entries",
        xaxis=dict(range=[-0.5, 9.5] if len(student_data_all) > 10 else None),
        showlegend=False
    )
    student_chart_html = student_chart.to_html(full_html=False)

    # Prepare years data
    years = sorted(data['scan_time'].dt.year.dropna().unique(), reverse=True)

    # Return context for rendering the template
    context = {
        'title_suffix': f" ({title_suffix})" if filters else '',
        'year_level_chart': year_level_chart_html,
        'section_chart': section_chart_html,
        'student_chart': student_chart_html,
        'year_levels': ['7', '8', '9', '10', '11', '12'],
        'selected_year_level': selected_year_level,
        'start_date': start_date,
        'end_date': end_date,
        'address': address_filter,
        'sex': sex_filter,
        'years': years
    }

    return render(request, 'logs/chart_reports.html', context)



def home(request):
    return redirect('login')  # Redirect root URL to the login page


def logout_view(request):
    logout(request)  # Logs the user out
    return redirect('login')  # Redirects the user to the login page after logout

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('index')  # Redirect to the main page after login
        else:
            messages.error(request, 'Invalid username or password')

    return render(request, 'login.html')  # Render login page

@login_required
@group_required('teacher')
def edit_student(request, student_id):
    student = get_object_or_404(Student, pk=student_id)

    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            print(request.POST)  # Log the POST data to confirm if 'sex' is included
            form.save()
            return redirect('students')
    else:
        form = StudentForm(instance=student)

    return render(request, 'logs/edit_student.html', {'form': form, 'student': student})

@login_required
@group_required('teacher')
def view_students(request):
    # Get search query and filters from the request
    search_query = request.GET.get('search', '')
    year_level_filter = request.GET.get('year_level', 'All')
    address_filter = request.GET.get('address', '')
    sex_filter = request.GET.get('sex', '')

    # Start with all students
    students = Student.objects.all()

    # Apply search filter if applicable
    if search_query:
        students = students.filter(
            Q(first_name__icontains=search_query) |
            Q(middle_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(learner_reference_number__icontains=search_query)
        )

    # Apply year level filter if applicable
    if year_level_filter and year_level_filter != 'All':
        students = students.filter(year_level=year_level_filter)

    # Apply address filter if applicable
    if address_filter:
        students = students.filter(address__icontains=address_filter)

    # Apply sex filter if applicable
    if sex_filter:
        students = students.filter(sex=sex_filter)

    # Count total students after applying filters
    total_students = students.count()

    # Pass the filtered students and the total student count to the template
    return render(request, 'logs/students.html', {
        'students': students,
        'total_students': total_students,  # Pass the total count
        'search_query': search_query,
        'year_level_filter': year_level_filter,
        'address_filter': address_filter,
        'sex_filter': sex_filter,
    })


@login_required
def index(request):
    return render(request, 'logs/index.html')  # This will render the index.html file

@login_required
@group_required('teacher')
def register_student(request):
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            print(request.POST)  # This will print the POST data to check if 'sex' is being captured correctly
            form.save()
            return redirect('students')
    else:
        form = StudentForm()

    return render(request, 'logs/register_student.html', {'form': form})

from django.views.decorators.csrf import csrf_exempt
import logging

logger = logging.getLogger(__name__)

from django.utils.timezone import localtime



@csrf_exempt
def scan_qr(request):
    if request.method == "POST":
        qr_data = request.POST.get("qr_data")
        print("Received from mobile:", qr_data)

        if not qr_data:
            return JsonResponse({'message': 'Missing QR data.'}, status=400)

        try:
            # Parse the scanned QR data
            first_name, middle_name, last_name, year_level, section = qr_data.split(',')
            first_name = first_name.strip()
            middle_name = middle_name.strip()
            last_name = last_name.strip()
            year_level = int(year_level.strip())
            section = section.strip()

            # Check if the student exists
            student = Student.objects.filter(
                first_name=first_name,
                middle_name=middle_name,
                last_name=last_name
            ).first()

            if not student:
                # Create new student if not found
                student = Student.objects.create(
                    first_name=first_name,
                    middle_name=middle_name,
                    last_name=last_name,
                    year_level=year_level,
                    section=section
                )

            today = timezone.localdate()
            if LateLog.objects.filter(student=student, scan_time__date=today).exists():
                return JsonResponse({'message': 'This QR code has already been scanned today.'})

            now = timezone.now()
            expected_time = now.replace(hour=7, minute=30, second=0, microsecond=0)
            late_minutes = max(0, int((now - expected_time).total_seconds() // 60))

            LateLog.objects.create(
                student=student,
                scan_time=now,
                late_minutes=late_minutes
            )

            # Convert 'now' to local time using localtime
            local_now = localtime(now)

            return JsonResponse({
                'message': f"Late entry recorded for {student.first_name} {student.last_name} at {local_now.strftime('%I:%M %p')}.",
                'late_minutes': late_minutes
            })

        except Exception as e:
            return JsonResponse({'message': f'Error: {str(e)}'}, status=400)

    # For GET requests, render the scanner page
    return render(request, 'logs/scan_qr.html')

from .models import Student

from datetime import datetime
from django.utils import timezone



def view_reports(request):
    # Get filter values from the GET request
    search = request.GET.get('search', '')
    year_level = request.GET.get('year_level', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    address = request.GET.get('address', '')
    sex = request.GET.get('sex', '')

    # Query LateLog model and filter based on the parameters
    logs = LateLog.objects.select_related('student').order_by('-scan_time')

    if search:
        logs = logs.filter(
            Q(student__first_name__icontains=search) |
            Q(student__last_name__icontains=search) |
            Q(student__middle_name__icontains=search)
        )

    if year_level:
        logs = logs.filter(student__year_level=year_level)

    if address:
        logs = logs.filter(student__address__icontains=address)

    if sex:
        logs = logs.filter(student__sex=sex)

    # Handle date range filtering
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            logs = logs.filter(scan_time__date__gte=start_date, scan_time__date__lte=end_date)
        except ValueError:
            pass  # If the date is invalid, no filter is applied

    # Convert scan_time to local time for display
    for log in logs:
        log.scan_time = localtime(log.scan_time)

    # Get today's late logs count
    today_count = LateLog.objects.filter(scan_time__date=timezone.localdate()).count()

    year_levels = ['7', '8', '9', '10', '11', '12']  # Example, adjust as needed

    return render(request, 'logs/view_reports.html', {
        'logs': logs,
        'search': search,
        'year_level': year_level,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'address': address,
        'sex': sex,
        'year_levels': year_levels,
        'today_count': today_count  # Pass count to template
    })



@login_required
@group_required('teacher')
def student_list(request):
    search_query = request.GET.get('search', '')
    year_level_filter = request.GET.get('year_level', '')
    address_filter = request.GET.get('address', '')
    sex_filter = request.GET.get('sex', '')

    # Start with all students
    students = Student.objects.all()

    # Apply search filter if applicable
    if search_query:
        students = students.filter(
            Q(first_name__icontains=search_query) |
            Q(middle_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(learner_reference_number__icontains=search_query)
        )

    # Apply year level filter if applicable
    if year_level_filter and year_level_filter != 'All':
        students = students.filter(year_level=year_level_filter)

    # Apply address filter if applicable
    if address_filter:
        students = students.filter(address__icontains=address_filter)

    # Apply sex filter if applicable
    if sex_filter:
        students = students.filter(sex=sex_filter)

    # Count the total number of students after applying filters
    total_students = students.count()

    # Export to Excel logic
    if 'export' in request.GET:
        df = pd.DataFrame(list(students.values(
            'learner_reference_number',
            'last_name',
            'first_name',
            'middle_name',
            'year_level',
            'section'
        )))
        df.rename(columns={
            'learner_reference_number': 'LRN',
            'last_name': 'Last Name',
            'first_name': 'First Name',
            'middle_name': 'Middle Name',
            'year_level': 'Year Level',
            'section': 'Section',
        }, inplace=True)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="student_list.xlsx"'
        df.to_excel(response, index=False)
        return response

    return render(request, 'logs/students.html', {
        'students': students,
        'year_levels': Student.objects.values_list('year_level', flat=True).distinct(),
        'selected_year': year_level_filter,
        'search_query': search_query,
        'address_filter': address_filter,
        'sex_filter': sex_filter,
        'total_students': total_students,  # Pass the total number of students to the template
    })


def export_report(request):
    # Get filter values from the GET request
    search = request.GET.get('search', '')
    year_level = request.GET.get('year_level', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    address = request.GET.get('address', '')
    sex = request.GET.get('sex', '')

    # Define thin border
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Fetch the logs based on filters
    logs = LateLog.objects.select_related('student').order_by('scan_time')

    if search:
        logs = logs.filter(
            Q(student__first_name__icontains=search) |
            Q(student__last_name__icontains=search) |
            Q(student__middle_name__icontains=search)
        )

    if year_level:
        logs = logs.filter(student__year_level=year_level)

    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            logs = logs.filter(scan_time__date__gte=start_date, scan_time__date__lte=end_date)
        except ValueError:
            pass

    if address:
        logs = logs.filter(student__address__icontains=address)

    if sex:
        logs = logs.filter(student__sex=sex)

    # Create the response and Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Late Logs Report"

    # Row 1: School name
    ws.merge_cells('A1:G1')
    ws['A1'].value = "CLAVER NATIONAL HIGH SCHOOL"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Row 2: Blank
    ws['A2'] = ''

    # Row 3: Report title
    ws.merge_cells('A3:G3')
    ws['A3'].value = "Late Logs Report"
    ws['A3'].font = Font(bold=True, size=14)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    # Row 4: Dynamic Title based on Filters (without parentheses)
    title_parts = []
    
    if year_level:
        title_parts.append(f"Grade {year_level}")

    if start_date_str and end_date_str:
        title_parts.append(f"From {datetime.strptime(start_date_str, '%Y-%m-%d').strftime('%m/%d/%Y')} To {datetime.strptime(end_date_str, '%Y-%m-%d').strftime('%m/%d/%Y')}")

    if address:
        title_parts.append(f"Brgy. {address}")

    if sex == 'M':
        title_parts.append("Male")
    elif sex == 'F':
        title_parts.append("Female")

    title_suffix = ', '.join(title_parts)

    # Merge title in row 4
    ws.merge_cells('A4:G4')
    ws['A4'].value = title_suffix  # Removed the parentheses here
    ws['A4'].font = Font(bold=True, size=14)
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')

    # Row 5: Blank
    ws['A5'] = ''

    # Row 6: Headers
    headers = ['Learner Reference Number', 'Student Name', 'Address', 'Sex', 'Year Level', 'Section', 'Date', 'Time']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="007bff")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Fill data starting from row 7
    for idx, log in enumerate(logs, start=7):
        student = log.student
        student_name = f"{student.last_name}, {student.first_name} {student.middle_name or ''}".strip()
        address = student.address
        sex = 'Male' if student.sex == 'M' else 'Female'
        year_level = student.year_level
        section = student.section
        scan_time = localtime(log.scan_time).replace(tzinfo=None)
        date_str = scan_time.strftime("%B %d, %Y")
        time_str = scan_time.strftime("%I:%M %p")

        row_data = [
            log.student.learner_reference_number,
            student_name,
            address,
            sex,
            year_level,
            section,
            date_str,
            time_str
        ]

        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=idx, column=col_num, value=value)
            cell.border = thin_border
            # Center align Year Level, Section, Date, and Time columns
            if col_num in [5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal='center')

    total_entries = logs.count()

    # Add a blank row after data
    footer_start_row = logs.count() + 8  # 7 data start + count + 1 blank

    # Add footer row with total late entries, merged across columns E and F (5 & 6)
    ws.cell(row=footer_start_row, column=5).value = f"Total Late Entries: {total_entries}"
    footer_cell = ws.cell(row=footer_start_row, column=5)
    footer_cell.font = Font(bold=True)
    footer_cell.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=footer_start_row, start_column=5, end_row=footer_start_row, end_column=6)

    # Adjust column widths automatically
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except:
                    pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Generate the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=late_logs.xlsx'

    wb.save(response)
    return response


@login_required
@group_required('teacher')
def export_student_list(request):
    # Get filter values from the GET request
    search_query = request.GET.get('search', '')
    year_level_filter = request.GET.get('year_level', '')
    address_filter = request.GET.get('address', '')
    sex_filter = request.GET.get('sex', '')

    # Start with all students
    students = Student.objects.all()

    # Apply search filter if applicable
    if search_query:
        students = students.filter(
            Q(first_name__icontains=search_query) |
            Q(middle_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(learner_reference_number__icontains=search_query)
        )

    # Apply year level filter if applicable
    if year_level_filter and year_level_filter != 'All':
        students = students.filter(year_level=year_level_filter)

    # Apply address filter if applicable
    if address_filter:
        students = students.filter(address__icontains=address_filter)

    # Apply sex filter if applicable
    if sex_filter:
        students = students.filter(sex=sex_filter)

    # Order the students
    students = students.order_by('last_name', 'first_name', 'middle_name')

    # Export to Excel logic
    df = pd.DataFrame(list(students.values(
        'learner_reference_number',
        'last_name',
        'first_name',
        'middle_name',
        'year_level',
        'section',
        'address',
        'sex'
    )))

    # Rename columns for cleaner export
    df.rename(columns={
        'learner_reference_number': 'LRN',
        'last_name': 'Last Name',
        'first_name': 'First Name',
        'middle_name': 'Middle Name',
        'year_level': 'Year Level',
        'section': 'Section',
        'address': 'Address',
        'sex': 'Sex',
    }, inplace=True)

    # Merge First, Middle, and Last Name into a single "Student Name" column
    df['Student Name'] = df['Last Name'] + ', ' + df['First Name'] + ' ' + df['Middle Name'].fillna('')

    # Drop the separate 'First Name', 'Middle Name', and 'Last Name' columns
    df = df.drop(columns=['First Name', 'Middle Name', 'Last Name'])

    # Reorder columns so "Student Name" comes first
    df = df[['LRN', 'Student Name', 'Address', 'Sex', 'Year Level', 'Section']]

    # Create the response for the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="student_list.xlsx"'

    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Student List"

    # Style the header row
    headers = ['LRN', 'Student Name', 'Address', 'Sex', 'Year Level', 'Section']
    
    # Row 1: "CLAVER NATIONAL HIGH SCHOOL"
    ws.merge_cells('A1:F1')
    ws['A1'] = "CLAVER NATIONAL HIGH SCHOOL"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Row 2: Blank
    ws['A2'] = ''
    
    # Row 3: "Student List" title
    ws.merge_cells('A3:F3')
    ws['A3'] = "Student List"
    ws['A3'].font = Font(bold=True, size=14)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    # Row 4: Blank
    ws['A4'] = ''

    # Row 5: Headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="007bff")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    # Fill the data starting from row 6
    for idx, student in enumerate(students, start=6):
        row_data = [
            student.learner_reference_number,
            f"{student.last_name}, {student.first_name} {student.middle_name or ''}",
            student.address,
            'Male' if student.sex == 'M' else 'Female',
            student.year_level,
            student.section
        ]
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=idx, column=col_num, value=value)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            # Center align Year Level and Section columns
            if col_num in [5, 6]:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add the footer row for total students
    total_students_row = len(students) + 7
    ws.merge_cells(start_row=total_students_row, start_column=5, end_row=total_students_row, end_column=6)
    footer_cell = ws.cell(row=total_students_row, column=5, value=f"Total Registered Students: {len(students)}")
    footer_cell.font = Font(bold=True)
    footer_cell.alignment = Alignment(horizontal='center')

    # Adjust column widths to fit the content
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except:
                    pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Save the workbook to the response
    wb.save(response)
    return response
